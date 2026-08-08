from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

from app.config import EXPORT_DIR
from app.database import get_connection, rows_to_dicts
from app.schemas import ReportFilters


def resolve_date_range(filters: ReportFilters) -> tuple[date, date]:
    today = date.today()
    if filters.preset == "today":
        return today, today
    if filters.preset == "this_month":
        return today.replace(day=1), today
    if filters.preset == "last_month":
        first_this_month = today.replace(day=1)
        last_month_end = first_this_month.fromordinal(first_this_month.toordinal() - 1)
        return last_month_end.replace(day=1), last_month_end
    if filters.preset == "this_year":
        return date(today.year, 1, 1), today
    if filters.preset == "last_year":
        year = today.year - 1
        return date(year, 1, 1), date(year, 12, 31)
    if not filters.start_date or not filters.end_date:
        raise ValueError("Custom reports require start_date and end_date.")
    if filters.start_date > filters.end_date:
        raise ValueError("start_date cannot be after end_date.")
    return filters.start_date, filters.end_date


def _where_for_filters(filters: ReportFilters, alias: str = "s") -> tuple[str, list[str]]:
    clauses: list[str] = []
    values: list[str] = []
    for field in ("department", "program", "academic_year", "semester", "section"):
        value = getattr(filters, field)
        if value:
            clauses.append(f"{alias}.{field} = ?")
            values.append(value)
    return (" AND ".join(clauses), values)


def build_attendance_report(filters: ReportFilters) -> dict:
    start_date, end_date = resolve_date_range(filters)
    session_clauses = ["attendance_date BETWEEN ? AND ?"]
    session_values: list[str] = [start_date.isoformat(), end_date.isoformat()]
    for field in ("department", "program", "academic_year", "semester", "section"):
        value = getattr(filters, field)
        if value:
            session_clauses.append(f"{field} = ?")
            session_values.append(value)

    with get_connection() as conn:
        sessions = rows_to_dicts(
            conn.execute(
                f"""
                SELECT *
                FROM attendance_sessions
                WHERE {' AND '.join(session_clauses)}
                ORDER BY attendance_date DESC, id DESC
                """,
                session_values,
            ).fetchall()
        )

        rows: list[dict] = []
        for session in sessions:
            student_clauses = ["status = 'active'"]
            student_values: list[str] = []
            for field in ("department", "program", "academic_year", "semester", "section"):
                report_value = getattr(filters, field)
                session_value = session[field]
                value = report_value or session_value
                if value:
                    student_clauses.append(f"{field} = ?")
                    student_values.append(value)

            students = rows_to_dicts(
                conn.execute(
                    f"""
                    SELECT id, roll_number, full_name, department, program, academic_year, semester, section
                    FROM college_students
                    WHERE {' AND '.join(student_clauses)}
                    ORDER BY roll_number COLLATE NOCASE
                    """,
                    student_values,
                ).fetchall()
            )

            present_rows = conn.execute(
                """
                SELECT student_id, confidence, marked_at
                FROM attendance_records
                WHERE session_id = ?
                """,
                (session["id"],),
            ).fetchall()
            present_map = {row["student_id"]: row for row in present_rows}

            for student in students:
                record = present_map.get(student["id"])
                rows.append(
                    {
                        "date": session["attendance_date"],
                        "session": session["title"],
                        "roll_number": student["roll_number"],
                        "full_name": student["full_name"],
                        "department": student["department"],
                        "program": student["program"],
                        "academic_year": student["academic_year"],
                        "semester": student["semester"],
                        "section": student["section"],
                        "status": "Present" if record else "Absent",
                        "confidence": round(float(record["confidence"]), 2) if record else "",
                        "marked_at": record["marked_at"] if record else "",
                    }
                )

    present = sum(1 for row in rows if row["status"] == "Present")
    absent = sum(1 for row in rows if row["status"] == "Absent")
    return {
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "sessions": sessions,
        "rows": rows,
        "summary": {
            "sessions": len(sessions),
            "students": len({row["roll_number"] for row in rows}),
            "present": present,
            "absent": absent,
            "total": len(rows),
        },
    }


def export_attendance_report(filters: ReportFilters) -> Path:
    report = build_attendance_report(filters)
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    headers = [
        "Date",
        "Session",
        "Roll Number",
        "Full Name",
        "Department",
        "Program",
        "Academic Year",
        "Semester",
        "Section",
        "Status",
        "Confidence",
        "Marked At",
    ]
    ws.append(headers)

    for row in report["rows"]:
        ws.append(
            [
                row["date"],
                row["session"],
                row["roll_number"],
                row["full_name"],
                row["department"],
                row["program"],
                row["academic_year"],
                row["semester"],
                row["section"],
                row["status"],
                row["confidence"],
                row["marked_at"],
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F2937")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    widths = [14, 24, 16, 26, 18, 18, 16, 12, 10, 12, 12, 22]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    filename = f"attendance_{report['start_date']}_to_{report['end_date']}.xlsx"
    output_path = EXPORT_DIR / filename
    wb.save(output_path)
    return output_path
